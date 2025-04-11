import openpyxl

wb = openpyxl.load_workbook("data.xlsx")


sheets = wb.sheetnames
wb.active.title
#print(wb.active.title)
sh1 = wb['Names']
data = sh1['B2'].value
data2 = wb['Names']['A2'].value
print(data2)

#alt alternative option
print(sh1.cell(3,2).value)
print(sh1.cell(3,3).value)
sh2=wb['Marks']

print(sh2.cell(2,1).value)
print(sh2.cell(3,2).value)

#alt alt 2

print(sh2.cell(row=2,column=2).value)
row = sh1.max_row
column = sh1.max_column

for i in range(1,row+1):
    for j in range(1, column+1):
        print(sh1.cell(i,j).value)


sh1.cell(row=5, column=1,value="Pytest")
sh1.cell(row=5, column=2, value = "UK")
sh1.cell(row=5, column=3, value= 88.88)
wb.save('Report.xlsx') #creae new sheet with old and new data

